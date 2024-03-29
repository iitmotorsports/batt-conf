from math import sqrt
import os
import sys
from typing import List
import importlib
import platform

if list(map(int, platform.python_version_tuple())) >= [3, 11, 0]:
    sys.exit("Current libraries only support below python 3.11")

try:
    import openpyxl
    import pandas
    import cadquery as cq
    import tqdm
except ImportError:
    print("Some required packages are missing. Installing now...")
    os.system('pip3.10 install -r requirements.txt')
    importlib.invalidate_caches()

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Color
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

from cell import CELLS
from calc import config_t

# Range Check
RANGE_SEGMENT_SERIES = list(range(1, 25))
RANGE_SEGMENT_PARALLEL = list(range(1, 65))
RANGE_ACCUMULATOR_SEGMENTS_SERIES = list(range(1, 13))
RANGE_ACCUMULATOR_SEGMENTS_PARALLEL = list(range(1, 13))

RANGES = (
    RANGE_SEGMENT_SERIES,
    RANGE_SEGMENT_PARALLEL,
    RANGE_ACCUMULATOR_SEGMENTS_SERIES,
    RANGE_ACCUMULATOR_SEGMENTS_PARALLEL,
)

# Targets
VOLT_TARGET = 450
VOLT_MAX = 600
VOLT_MIN = 400
CAP_TOTAL_MAX_WH = 10000
CAP_TOTAL_MIN_WH = 2000
CAP_SEGMENT_MAX_MJ = 6.1
VOLT_SEGMENT_MAX = 120.1
WEIGHT_LIMIT_KG = 60
POWER_NOM_MIN_KW = 40
POWER_MAX_MIN_KW = 70

ALLOW_ODD_SERIES_SEGMENTS = False
ALLOW_ODD_SERIES_CELLS = True

# ADDITIONAL_CONFIG = [
#     (6, 1, 16, 8),
# ]

favorLarge = ColorScaleRule(
    start_type='min', start_color=Color(rgb='ea0c01'),
    mid_type='percentile', mid_value=50, mid_color=Color(rgb='fef301'),
    end_type='max', end_color=Color(rgb='0ba60a')
)
favorSmall = ColorScaleRule(
    start_type='min', start_color=Color(rgb='0ba60a'),
    mid_type='percentile', mid_value=50, mid_color=Color(rgb='fef301'),
    end_type='max', end_color=Color(rgb='ea0c01')
)
suggestLarge = ColorScaleRule(
    start_type='min', start_color=Color(rgb='ffffff'),
    end_type='max', end_color=Color(rgb='006796')
)
suggestSmall = ColorScaleRule(
    start_type='min', start_color=Color(rgb='006796'),
    end_type='max', end_color=Color(rgb='ffffff')
)

conditionals = {
    "V∆": suggestLarge,
    "Price": favorSmall,
    "Kg": favorSmall,
    "m³": favorSmall,
    "Power Nom kW": favorLarge,
    "Power Max kW": favorLarge,
    "Typ Cap Wh": favorLarge,
    "Max Cap Wh": favorLarge,
    "Wh/Kg": favorLarge,
    "kW/Kg Max": favorLarge,
    "kW/Kg Nom": favorLarge,
    "Discharge Cont. A": favorLarge,
    "Discharge Peak A": favorLarge,
}

accending_sort = [
    ('Price', True),
    ('Kg', True),
    ('Max Cap Wh', False),
    ('Power Max kW', False),
    ('Typ Cap Wh', False),
    ('Power Nom kW', False),
    ('Power Nom kW', False),
    ('Wh/Kg', False),
    ('kW/Kg Max', False),
    ('kW/Kg Nom', False),
]

# for ss, sp, s, p in ADDITIONAL_CONFIG:
#     RANGE_SEGMENT_SERIES.append(ss)
#     RANGE_SEGMENT_PARALLEL.append(sp)
#     RANGE_ACCUMULATOR_SEGMENTS_SERIES.append(s)
#     RANGE_ACCUMULATOR_SEGMENTS_PARALLEL.append(p)

results: List[config_t] = []

for cell_sel in CELLS:
    for pk_parallel in RANGE_ACCUMULATOR_SEGMENTS_PARALLEL:
        for pk_series in RANGE_ACCUMULATOR_SEGMENTS_SERIES:
            if not ALLOW_ODD_SERIES_SEGMENTS and pk_series != 1 and pk_series % 2 != 0:
                continue
            for c_parallel in RANGE_SEGMENT_PARALLEL:
                for c_series in RANGE_SEGMENT_SERIES:
                    if not ALLOW_ODD_SERIES_CELLS and c_series != 1 and c_series % 2 != 0:
                        continue
                    conf = config_t(cell_sel, pk_parallel, pk_series, c_parallel, c_series, VOLT_TARGET)
                    if VOLT_MIN <= conf.total_volt_nominal <= VOLT_MAX and CAP_TOTAL_MIN_WH <= conf.total_typ_capacity_wh <= CAP_TOTAL_MAX_WH:
                        if conf.segment_volt_max <= VOLT_SEGMENT_MAX and conf.segment_typ_capacity_mj <= CAP_SEGMENT_MAX_MJ:
                            if conf.weight_kg <= WEIGHT_LIMIT_KG and conf.power_nom_kw >= POWER_NOM_MIN_KW and conf.power_max_kw >= POWER_MAX_MIN_KW:
                                results.append(conf)

df = pd.DataFrame([[y for y in x.info().values()] for x in results], columns=list(results[0].info().keys()))
for k, v in accending_sort:
    df.sort_values(k, ascending=v, inplace=True)


wb = load_workbook(filename="VBA.xlsm", keep_vba=True)
ws = wb.active

for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

ws.insert_cols(2)
ws.cell(row=1, column=2, value="CAD")
font = Font(underline="single", color="006796")
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=2, value="Generate")
    cell.font = font

ws.auto_filter.ref = ws.dimensions

for i, cell in enumerate(ws[1], start=1):
    if cell.value in conditionals:
        letter = get_column_letter(i)
        ws.conditional_formatting.add(f"{letter}2:{letter}{len(results)*2}", conditionals[cell.value])  # IMPROVE: extend to the entire column

for column in ws.columns:
    max_length = 0
    column_dimensions = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except TypeError:
            pass
    adjusted_width = max_length + (10 / sqrt(max_length+1))
    ws.column_dimensions[column_dimensions].width = adjusted_width

wb.save('output.xlsm')
os.system('output.xlsm')
